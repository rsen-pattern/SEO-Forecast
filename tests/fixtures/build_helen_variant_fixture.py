"""Builds sample_helen_variant_roadmap.xlsx mirroring the real Helen Kaminski layout.

Key structural differences from sample_pattern_native_roadmap.xlsx:
- Blank col A across all sheets (labels are in col B, values in col C)
- Client Detail has "Monthly Retainer (excl tech fees)" label and tooltip in col D
- Consulting rows use "6 Months" cadence
- Technical has "Month 1" string values in col A
- Content has primary-domain and localisation rows
- Links rows are present

Run directly:
    python tests/fixtures/build_helen_variant_fixture.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

FIXTURE_DIR = Path(__file__).parent
FIXTURE_PATH = FIXTURE_DIR / "sample_helen_variant_roadmap.xlsx"


def build(output_path: str | Path | None = None) -> Path:
    out = Path(output_path) if output_path else FIXTURE_PATH
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Breakdown ─────────────────────────────────────────────────────────────
    ws_bd = wb.create_sheet("Breakdown")
    ws_bd.cell(row=1, column=1, value="SEO Retainer Breakdown")

    def _set_breakdown_row(row, label, hours_list):
        ws_bd.cell(row=row, column=5, value=label)
        for i, h in enumerate(hours_list):
            ws_bd.cell(row=row, column=7 + i, value=h)

    _set_breakdown_row(4, "Consulting Hours", [8] * 12)
    _set_breakdown_row(6, "Technical Hours", [12] * 12)
    _set_breakdown_row(8, "Content Hours", [20] * 12)
    _set_breakdown_row(10, "Link Hours", [6] * 12)

    # ── Client Detail — blank col A, labels in col B, values in col C ─────────
    ws_cd = wb.create_sheet("1. Client Detail")
    ws_cd.cell(row=1, column=2, value="SEO Retainer Agreement")

    # Col A intentionally blank on all rows to simulate Helen file layout
    client_data = [
        # (row, col_B_label, col_C_value, col_D_tooltip)
        (3, "Client Name", "Helen Kaminski", None),
        (4, "Industry", "Accessories", None),
        (5, "Monthly Retainer (excl tech fees)", 4906, "(excl. tech — see separate line)"),
        (6, "Project Start Date", "2026-01-01", None),
        (7, "CMS", "Shopify", None),
        (8, "Primary Contact", "to be provided", "Please enter contact name"),
        (9, "Account Manager", "SEO to add", None),
        (10, "Domain", "double click to enter", None),
        (11, "Notes", "TBC: confirm retainer split", None),
    ]
    for row, label, value, tooltip in client_data:
        ws_cd.cell(row=row, column=2, value=label)
        ws_cd.cell(row=row, column=3, value=value)
        if tooltip:
            ws_cd.cell(row=row, column=4, value=tooltip)

    # ── Consulting — blank col A, header in row 2, "6 Months" cadence ─────────
    ws_con = wb.create_sheet("2. Consulting")
    ws_con.cell(row=1, column=2, value="Consulting Tasks")

    headers = ["Task", "Focus", "Occurrence", "Hours"]
    for ci, h in enumerate(headers, 2):
        ws_con.cell(row=2, column=ci, value=h)

    consulting_tasks = [
        ("Monthly Strategy Review", "Strategy", "Monthly", 4),
        ("Keyword Research & Mapping", "Strategy", "6 Months", 8),
        ("Competitor Analysis", "Strategy", "6 Months", 6),
        ("GA4 & GSC Reporting", "Analytics", "Monthly", 3),
        ("Content Gap Analysis", "Strategy", "6 Months", 8),
        ("Backlink Profile Review", "Off-Page", "6 Months", 4),
    ]
    for ri, (task, focus, occ, hrs) in enumerate(consulting_tasks, 3):
        ws_con.cell(row=ri, column=2, value=task)
        ws_con.cell(row=ri, column=3, value=focus)
        ws_con.cell(row=ri, column=4, value=occ)
        ws_con.cell(row=ri, column=5, value=hrs)

    # ── Technical — col A has "Month 1" / "Month 2" strings ──────────────────
    ws_tech = wb.create_sheet("3. Technical")
    ws_tech.cell(row=1, column=2, value="Technical SEO Tasks")

    headers = ["Task", "Focus", "Occurrence", "Hours"]
    for ci, h in enumerate(headers, 2):
        ws_tech.cell(row=2, column=ci, value=h)

    technical_tasks = [
        ("Month 1", "Core Web Vitals Audit", "Technical", "Monthly", 6),
        ("Month 1", "Crawl Error Remediation", "Technical", "Monthly", 4),
        ("Month 2", "Schema Markup Implementation", "Technical", "6 Months", 8),
        ("Month 2", "Site Speed Optimisation", "Technical", "Monthly", 4),
        ("Month 3", "Internal Linking Audit", "Technical", "6 Months", 10),
        ("Month 3", "Redirect Chain Cleanup", "Technical", "6 Months", 6),
    ]
    for ri, (month_str, task, focus, occ, hrs) in enumerate(technical_tasks, 3):
        ws_tech.cell(row=ri, column=1, value=month_str)
        ws_tech.cell(row=ri, column=2, value=task)
        ws_tech.cell(row=ri, column=3, value=focus)
        ws_tech.cell(row=ri, column=4, value=occ)
        ws_tech.cell(row=ri, column=5, value=hrs)

    # ── Content — primary domain + localisation rows ──────────────────────────
    ws_cont = wb.create_sheet("4. Content")
    ws_cont.cell(row=1, column=2, value="Content Production Plan")

    content_headers = ["Month", "Month Name", "URL", "Title", "Focus", "Priority",
                       "Content Type", "Word Count", "SEO Hours", "Brief / Notes"]
    for ci, h in enumerate(content_headers, 2):
        ws_cont.cell(row=5, column=ci, value=h)

    primary = "helenkaminski.com"
    localisation_au = "helenkaminski.com.au"
    localisation_nz = "helenkaminski.co.nz"

    content_rows = [
        # (month, name, url, title, focus, priority, ctype, wc, hrs, brief)
        # Primary domain — 300 words, new pages
        (1, "Jan", f"https://{primary}/blog/silk-scarves-guide", "Silk Scarves Style Guide", "Content", "High", "New Page", 300, 2.0, ""),
        (1, "Jan", f"https://{primary}/blog/accessories-trends", "Accessories Trends 2026", "Content", "High", "New Page", 300, 2.0, ""),
        (2, "Feb", f"https://{primary}/blog/straw-bags-summer", "Straw Bags Summer Edit", "Content", "Medium", "New Page", 300, 2.0, ""),
        (2, "Feb", f"https://{primary}/blog/raffia-hat-guide", "Raffia Hat Ultimate Guide", "Content", "Medium", "New Page", 300, 2.0, ""),
        (3, "Mar", f"https://{primary}/blog/gift-guide-women", "Gift Guide for Women", "Content", "High", "New Page", 300, 2.0, ""),
        # Existing page optimisations — primary domain
        (1, "Jan", f"https://{primary}/hats", "Hats Category Page", "Content", "High", "Optimisation", 200, 1.0, ""),
        (2, "Feb", f"https://{primary}/scarves", "Scarves Category Page", "Content", "Medium", "Optimisation", 200, 1.0, ""),
        (3, "Mar", f"https://{primary}/bags", "Bags Category Page", "Content", "Medium", "Optimisation", 200, 1.0, ""),
        # FAQ pages
        (1, "Jan", f"https://{primary}/faq/hat-sizing", "Hat Sizing FAQ", "Content", "Low", "FAQ", 150, 1.0, ""),
        (2, "Feb", f"https://{primary}/faq/care-instructions", "Care Instructions FAQ", "Content", "Low", "FAQ", 150, 1.0, ""),
        # Localisation — AU — 150 words
        (1, "Jan", f"https://{localisation_au}/blog/silk-scarves-guide", "Silk Scarves Style Guide AU", "Content", "High", "New Page", 150, 1.5, "localisation"),
        (1, "Jan", f"https://{localisation_au}/blog/accessories-trends", "Accessories Trends 2026 AU", "Content", "High", "New Page", 150, 1.5, "localisation"),
        (2, "Feb", f"https://{localisation_au}/blog/straw-bags-summer", "Straw Bags Summer Edit AU", "Content", "Medium", "New Page", 150, 1.5, "localisation"),
        (3, "Mar", f"https://{localisation_au}/hats", "Hats AU Category Page", "Content", "High", "Optimisation", 100, 1.0, "localisation"),
        (3, "Mar", f"https://{localisation_au}/scarves", "Scarves AU Category Page", "Content", "Medium", "Optimisation", 100, 1.0, "localisation"),
        # Localisation — NZ
        (2, "Feb", f"https://{localisation_nz}/blog/straw-bags-summer", "Straw Bags Summer Edit NZ", "Content", "Low", "New Page", 150, 1.5, "localisation"),
        (3, "Mar", f"https://{localisation_nz}/hats", "Hats NZ Category Page", "Content", "Low", "Optimisation", 100, 1.0, "localisation"),
    ]
    for ri, row_data in enumerate(content_rows, 6):
        for ci, val in enumerate(row_data, 2):
            ws_cont.cell(row=ri, column=ci, value=val)

    # ── Links — blank col A ───────────────────────────────────────────────────
    ws_links = wb.create_sheet("5. Links")
    ws_links.cell(row=1, column=2, value="Link Building Tasks")

    for ci, h in enumerate(["Task", "Focus", "Occurrence", "Hours"], 2):
        ws_links.cell(row=2, column=ci, value=h)

    links_tasks = [
        ("Digital PR Outreach", "Off-Page", "Monthly", 4),
        ("Guest Post Sourcing", "Off-Page", "Monthly", 2),
        ("Link Reclamation", "Off-Page", "6 Months", 3),
        ("Competitor Backlink Analysis", "Strategy", "6 Months", 4),
    ]
    for ri, (task, focus, occ, hrs) in enumerate(links_tasks, 3):
        ws_links.cell(row=ri, column=2, value=task)
        ws_links.cell(row=ri, column=3, value=focus)
        ws_links.cell(row=ri, column=4, value=occ)
        ws_links.cell(row=ri, column=5, value=hrs)

    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"Fixture written to: {path}")
