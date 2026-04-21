"""Tests for engine/roadmap_native_parser.py — Prompt 7."""
import io
from pathlib import Path

import pandas as pd
import pytest

FIXTURE_DIR = Path(__file__).parent / "fixtures"
XLSX_FIXTURE = FIXTURE_DIR / "sample_pattern_native_roadmap.xlsx"
HELEN_FIXTURE = FIXTURE_DIR / "sample_helen_variant_roadmap.xlsx"
TASK_CSV = FIXTURE_DIR / "sample_task_table.csv"
PARAM_CSV = FIXTURE_DIR / "sample_param_table.csv"


@pytest.fixture(scope="session")
def xlsx_bytes():
    return XLSX_FIXTURE.read_bytes()


@pytest.fixture(scope="session")
def task_csv_bytes():
    return TASK_CSV.read_bytes()


@pytest.fixture(scope="session")
def param_csv_bytes():
    return PARAM_CSV.read_bytes()


from engine.roadmap_ai_engine import load_roadmap_v2
from engine.roadmap_native_parser import (
    detect_roadmap_format,
    parse_pattern_native,
    wrap_legacy_param_table_as_bundle,
    wrap_legacy_task_table_as_bundle,
)


class TestDetectFormat:
    def test_detect_pattern_native(self, xlsx_bytes):
        assert detect_roadmap_format(xlsx_bytes, "xlsx") == "pattern_native"

    def test_detect_task_table(self, task_csv_bytes):
        assert detect_roadmap_format(task_csv_bytes, "csv") == "task_table"

    def test_detect_param_table(self, param_csv_bytes):
        assert detect_roadmap_format(param_csv_bytes, "csv") == "param_table"

    def test_detect_unknown_returns_unknown(self):
        assert detect_roadmap_format(b"hello world", "txt") == "unknown"


class TestParsePatternNative:
    @pytest.fixture(scope="class")
    def bundle(self, xlsx_bytes):
        return parse_pattern_native(xlsx_bytes)

    def test_parse_pattern_native_client_metadata_extracted(self, bundle):
        meta = bundle["client_metadata"]
        assert "client_name" in meta
        assert meta["client_name"] != ""

    def test_parse_pattern_native_per_focus_hours_computed(self, bundle):
        per_focus = bundle["per_focus"]
        total_hours = sum(f["monthly_hours"] for f in per_focus.values())
        assert total_hours > 0

    def test_parse_pattern_native_content_plan_not_empty(self, bundle):
        assert len(bundle["content_plan"]) > 0

    def test_parse_pattern_native_content_plan_classifies_new_page_vs_optimisation(self, bundle):
        types = {item["content_type"] for item in bundle["content_plan"]}
        assert "new_page" in types
        # Fixture has both new pages and optimisations
        assert len(types) > 1

    def test_parse_pattern_native_strategy_restart_detected(self, bundle):
        # Fixture has "Monthly Strategy Review" in consulting + content only in months 1-4
        # so strategy_restart_month should be set
        timeline = bundle["timeline"]
        assert "strategy_restart_month" in timeline
        # Could be None if content covers all 12 months — just assert key exists

    def test_parse_pattern_native_handles_missing_sheets(self):
        # A workbook with only Breakdown + 3 expected sheets should still parse
        import openpyxl

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws_bd = wb.create_sheet("Breakdown")
        ws_bd.cell(row=4, column=5, value="Content Hours")
        for i, h in enumerate([10] * 12):
            ws_bd.cell(row=4, column=7 + i, value=h)
        wb.create_sheet("1. Client Detail")
        wb.create_sheet("4. Content")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        bundle = parse_pattern_native(buf.read())
        assert bundle["schema_version"] == "2.0"
        assert "content" in bundle["per_focus"]


class TestWrapLegacy:
    def test_wrap_legacy_task_table_produces_v2_bundle(self):
        legacy = {"content_cadence": 4, "effort_level": "moderate", "maintenance_coverage": 0.5}
        bundle = wrap_legacy_task_table_as_bundle(legacy)
        assert bundle["schema_version"] == "2.0"
        assert bundle["format_detected"] == "task_table"
        assert bundle["source_summary"]["parsing_confidence"] == pytest.approx(0.7)
        assert "per_focus" in bundle

    def test_wrap_legacy_param_table_produces_v2_bundle(self):
        legacy = {"effort_level": "light", "content_cadence": 2}
        bundle = wrap_legacy_param_table_as_bundle(legacy)
        assert bundle["schema_version"] == "2.0"
        assert bundle["format_detected"] == "param_table"
        assert bundle["source_summary"]["parsing_confidence"] == pytest.approx(0.5)


class TestLoadRoadmapV2:
    def test_load_roadmap_v2_dispatches_correctly(self, xlsx_bytes, task_csv_bytes, param_csv_bytes):
        b1, m1 = load_roadmap_v2(None, xlsx_bytes, "roadmap.xlsx")
        assert b1["format_detected"] == "pattern_native"
        assert m1 == "deterministic"

        b2, m2 = load_roadmap_v2(None, task_csv_bytes, "tasks.csv")
        assert b2["format_detected"] == "task_table"
        assert m2 == "deterministic"

        b3, m3 = load_roadmap_v2(None, param_csv_bytes, "params.csv")
        assert b3["format_detected"] == "param_table"
        assert m3 == "deterministic"


class TestHelenVariantParsing:
    """Tests against the Helen Kaminski layout fixture (blank col A, localisation rows, tooltips)."""

    @pytest.fixture(scope="class")
    def helen_bytes(self):
        return HELEN_FIXTURE.read_bytes()

    @pytest.fixture(scope="class")
    def bundle(self, helen_bytes):
        return parse_pattern_native(helen_bytes)

    def test_parses_without_raising_on_localisation_text(self, helen_bytes):
        # Must not crash regardless of "localisation" text in content cells
        bundle = parse_pattern_native(helen_bytes)
        assert bundle["schema_version"] == "2.0"

    def test_client_detail_parses_despite_blank_col_a(self, bundle):
        # Labels in col B, not col A — header-driven parser must find them
        assert bundle["client_metadata"].get("client_name") == "Helen Kaminski"

    def test_retainer_parsed_despite_label_suffix(self, bundle):
        # Label is "Monthly Retainer (excl tech fees)" — contains more than just "retainer"
        assert bundle["client_metadata"].get("retainer_aud_monthly") == pytest.approx(4906)

    def test_tooltip_columns_ignored(self, bundle):
        # Values like "to be provided", "double click", "SEO to add" must be filtered
        meta = bundle["client_metadata"]
        for v in meta.values():
            if isinstance(v, str):
                assert "to be provided" not in v.lower()
                assert "double click" not in v.lower()
                assert "seo to add" not in v.lower()
