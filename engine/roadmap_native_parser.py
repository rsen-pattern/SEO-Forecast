"""Deterministic parser for Pattern-native xlsx roadmap format.

Parses the multi-sheet xlsx produced by Pattern's retainer template without
requiring an AI call. Falls back to legacy scalar wrappers for simpler formats.

Pure Python — no Streamlit imports.
"""
from __future__ import annotations

import io
import re
from typing import Any

import openpyxl

# ── Sheet name constants ──────────────────────────────────────────────────────

PATTERN_NATIVE_SHEETS_REQUIRED = {"Breakdown"}
PATTERN_NATIVE_SHEETS_EXPECTED = {
    "1. Client Detail",
    "2. Consulting",
    "3. Technical",
    "4. Content",
    "5. Links",
}

# ── Normalisation maps ────────────────────────────────────────────────────────

INDUSTRY_NORMALISATION: dict[str, str] = {
    "accessories": "Accessories",
    "apparel": "Apparel",
    "clothing": "Apparel",
    "fashion": "Apparel",
    "beauty": "Beauty",
    "cosmetics": "Beauty",
    "home": "Home",
    "homewares": "Home",
    "b2b saas": "B2B SaaS",
    "saas": "B2B SaaS",
    "automotive": "Automotive",
    "travel": "Travel",
    "food": "Food & Beverage",
    "health": "Health",
    "fitness": "Health",
    "finance": "Finance",
    "fintech": "Finance",
}

# ── Effort classification thresholds (avg monthly hours) ─────────────────────

_EFFORT_HOURS_THRESHOLDS: list[tuple[float, str]] = [
    (8.0, "light"),
    (20.0, "moderate"),
    (float("inf"), "aggressive"),
]

# ── Column aliases for format detection ──────────────────────────────────────

_TASK_COL_ALIASES = {"task", "activity"}
_FOCUS_COL_ALIASES = {"focus", "area", "category"}
_OCC_COL_ALIASES = {"occurrence", "frequency"}
_HOURS_COL_ALIASES = {"hours", "hrs"}
_CADENCE_COL_ALIASES = {"cadence", "content_cadence", "posts_per_month"}
_EFFORT_COL_ALIASES = {"effort_level", "effort"}
_MAINT_COL_ALIASES = {"maintenance_coverage", "maintenance", "maint_coverage"}

_MONTH_PATTERN = re.compile(r"month\s*(\d+)", re.IGNORECASE)

_TOOLTIP_FRAGMENTS = (
    "to be provided",
    "double click",
    "seo to add",
    "tbc:",
    "need access",
)

# Occurrence string → monthly hours multiplier (longest-key match first)
_OCC_MULTIPLIERS: list[tuple[str, float]] = [
    ("monthly basis", 1.0),
    ("bi-monthly", 0.5),
    ("every 2 months", 0.5),
    ("quarterly", 0.33),
    ("every 3 months", 0.33),
    ("6 months", 0.167),
    ("bi-annual", 0.167),
    ("every 6 months", 0.167),
    ("half year", 0.167),
    ("monthly", 1.0),
    ("annual", 0.083),
    ("one-off", 0.083),
    ("once", 0.083),
]


# ── Safe conversion helpers ───────────────────────────────────────────────────


def _safe_int(val, default: int = 0) -> int:
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _safe_float(val, default: float = 0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _parse_month_string(val) -> int | None:
    """Extract month number from strings like 'Month 1', 'Month  2', 'Month1'."""
    if val is None:
        return None
    m = _MONTH_PATTERN.search(str(val))
    return int(m.group(1)) if m else None


def _find_header_row(ws, contains_all: set[str], scan_max: int = 15) -> int | None:
    """Return 1-indexed row number of first row where all strings appear (case-insensitive)."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_max, values_only=True), 1):
        row_text = " ".join(str(c).lower() for c in row if c is not None)
        if all(s.lower() in row_text for s in contains_all):
            return row_idx
    return None


def _ws_to_markdown(ws, max_chars: int = 20000) -> tuple[str, bool]:
    """Convert worksheet to a markdown table string for AI consumption."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append([str(v).strip() if v is not None else "" for v in row])

    if not rows:
        return "", False

    last_col = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i].strip():
                last_col = max(last_col, i)
                break

    col_count = last_col + 1
    rows = [r[:col_count] for r in rows]

    def _pad(r: list[str]) -> list[str]:
        return r + [""] * (col_count - len(r))

    lines = [
        "| " + " | ".join(_pad(rows[0])) + " |",
        "| " + " | ".join(["---"] * col_count) + " |",
    ] + ["| " + " | ".join(_pad(r)) + " |" for r in rows[1:]]

    md = "\n".join(lines)
    was_truncated = len(md) > max_chars
    return md[:max_chars], was_truncated


def _occ_to_multiplier(occurrence: str) -> float:
    s = occurrence.strip().lower()
    for key, mult in _OCC_MULTIPLIERS:
        if key in s:
            return mult
    return 1.0


# ── Format detection ──────────────────────────────────────────────────────────


def detect_roadmap_format(raw_bytes: bytes, file_extension: str) -> str:
    """Detect which roadmap format the bytes represent.

    Returns one of: "pattern_native", "task_table", "param_table", "unknown".
    """
    ext = file_extension.lower().lstrip(".")

    # Only xlsx can be Pattern-native
    if ext in ("xlsx", "xls"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
            sheet_names = set(wb.sheetnames)
            wb.close()
            if PATTERN_NATIVE_SHEETS_REQUIRED.issubset(sheet_names):
                matching_expected = sheet_names & PATTERN_NATIVE_SHEETS_EXPECTED
                if len(matching_expected) >= 3:
                    return "pattern_native"
        except Exception:
            pass

    # Try reading as a flat table (csv or xlsx)
    try:
        import pandas as pd

        buf = io.BytesIO(raw_bytes)
        if ext in ("xlsx", "xls"):
            df = pd.read_excel(buf, engine="openpyxl")
        else:
            try:
                df = pd.read_csv(buf)
            except Exception:
                return "unknown"

        cols_lower = {str(c).strip().lower() for c in df.columns}

        has_task = bool(cols_lower & _TASK_COL_ALIASES)
        has_focus = bool(cols_lower & _FOCUS_COL_ALIASES)
        has_occ = bool(cols_lower & _OCC_COL_ALIASES)
        has_hours = bool(cols_lower & _HOURS_COL_ALIASES)
        has_cadence = bool(cols_lower & _CADENCE_COL_ALIASES)
        has_effort = bool(cols_lower & _EFFORT_COL_ALIASES)
        has_maint = bool(cols_lower & _MAINT_COL_ALIASES)

        # param_table: has cadence or effort_level or maintenance_coverage
        if has_cadence or has_maint or (has_effort and not has_task and not has_focus):
            return "param_table"

        # task_table: has task + focus + occurrence + hours (at least 3 of 4)
        task_score = sum([has_task, has_focus, has_occ, has_hours])
        if task_score >= 3:
            return "task_table"

    except Exception:
        pass

    return "unknown"


# ── Empty per-focus skeleton ──────────────────────────────────────────────────


def _empty_per_focus() -> dict:
    """Return a per_focus dict with all 7 focus areas at zero/moderate defaults."""
    focus_areas = ["content", "technical", "on_page", "off_page", "local", "analytics", "strategy"]
    return {
        area: {
            "effort_level": "moderate",
            "monthly_hours": 0.0,
            "cadence": 0,
            "task_count": 0,
            "tasks": [],
        }
        for area in focus_areas
    }


# ── Client detail parser ──────────────────────────────────────────────────────


def _parse_client_detail(ws) -> dict:
    """Parse the '1. Client Detail' worksheet into a metadata dict.

    Header-driven: scans first 15 rows for the cell containing 'client name'
    to discover which column holds labels vs. values. Handles files where col A
    is blank (real Helen Kaminski layout) as well as files with labels in col A.
    Tooltip/placeholder values are filtered out.
    """
    result: dict[str, Any] = {}

    all_rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))

    label_col: int | None = None
    for row in all_rows[:15]:
        for col_idx, cell in enumerate(row):
            if cell is not None and "client name" in str(cell).strip().lower():
                label_col = col_idx
                break
        if label_col is not None:
            break

    if label_col is None:
        return result

    value_col = label_col + 1

    def _is_tooltip(val) -> bool:
        if val is None:
            return True
        s = str(val).strip().lower()
        return not s or any(frag in s for frag in _TOOLTIP_FRAGMENTS)

    for row in all_rows:
        if not row or len(row) <= label_col:
            continue
        label_cell = row[label_col]
        if label_cell is None:
            continue

        label = str(label_cell).strip().lower()
        value = row[value_col] if len(row) > value_col else None

        if _is_tooltip(value):
            continue

        if "client name" in label:
            result["client_name"] = str(value).strip()

        elif "industry" in label:
            raw = str(value).strip().lower()
            result["industry"] = INDUSTRY_NORMALISATION.get(raw, str(value).strip())

        elif "retainer" in label:
            num_str = re.sub(r"[^0-9.]", "", str(value))
            try:
                result["retainer_aud_monthly"] = float(num_str)
            except (ValueError, TypeError):
                result["retainer_aud_monthly"] = str(value).strip()

        elif "project start" in label:
            result["project_start_date"] = str(value).strip()

        elif label == "cms" or label.startswith("cms"):
            result["cms"] = str(value).strip()

    return result


# ── Breakdown parser ──────────────────────────────────────────────────────────


def _parse_breakdown(ws) -> dict:
    """Parse the 'Breakdown' worksheet to extract monthly hours per service type."""
    result: dict[str, list] = {
        "consulting": [0] * 12,
        "technical": [0] * 12,
        "content": [0] * 12,
        "link": [0] * 12,
    }

    label_map = {
        "consulting hours": "consulting",
        "technical hours": "technical",
        "content hours": "content",
        "link hours": "link",
    }

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 7:
            continue
        cell_e = row[4]  # col E (0-indexed index 4)
        if cell_e is None:
            continue

        label_lower = str(cell_e).strip().lower()
        matched_key = None
        for label_fragment, key in label_map.items():
            if label_fragment in label_lower:
                matched_key = key
                break

        if matched_key is None:
            continue

        # Extract cols G:R (0-indexed 6:18)
        monthly_values = []
        for val in row[6:18]:
            if val is not None and val != "":
                try:
                    monthly_values.append(float(val))
                except (ValueError, TypeError):
                    pass

        # Pad or trim to 12
        while len(monthly_values) < 12:
            monthly_values.append(0.0)
        result[matched_key] = monthly_values[:12]

    return result


# ── Effort classification ─────────────────────────────────────────────────────


def _classify_effort_hours(hours: float) -> str:
    """Classify average monthly hours into an effort level label."""
    for threshold, label in _EFFORT_HOURS_THRESHOLDS:
        if hours <= threshold:
            return label
    return "aggressive"


# ── Apply breakdown to bundle ─────────────────────────────────────────────────


def _apply_breakdown_to_bundle(bundle: dict, breakdown: dict) -> None:
    """Map breakdown monthly hours into per_focus entries."""
    per_focus = bundle["per_focus"]

    def _avg_nonzero(values: list) -> float:
        nonzero = [v for v in values if v and v != 0.0]
        return sum(nonzero) / len(nonzero) if nonzero else 0.0

    # Consulting hours → 70% strategy, 30% analytics
    consulting_avg = _avg_nonzero(breakdown.get("consulting", []))
    strategy_hours = consulting_avg * 0.70
    analytics_hours = consulting_avg * 0.30
    per_focus["strategy"]["monthly_hours"] = round(strategy_hours, 2)
    per_focus["strategy"]["effort_level"] = _classify_effort_hours(strategy_hours)
    per_focus["analytics"]["monthly_hours"] = round(analytics_hours, 2)
    per_focus["analytics"]["effort_level"] = _classify_effort_hours(analytics_hours)

    # Technical hours → 70% technical, 30% on_page
    technical_avg = _avg_nonzero(breakdown.get("technical", []))
    tech_hours = technical_avg * 0.70
    on_page_hours = technical_avg * 0.30
    per_focus["technical"]["monthly_hours"] = round(tech_hours, 2)
    per_focus["technical"]["effort_level"] = _classify_effort_hours(tech_hours)
    per_focus["on_page"]["monthly_hours"] = round(on_page_hours, 2)
    per_focus["on_page"]["effort_level"] = _classify_effort_hours(on_page_hours)

    # Content hours → content
    content_avg = _avg_nonzero(breakdown.get("content", []))
    per_focus["content"]["monthly_hours"] = round(content_avg, 2)
    per_focus["content"]["effort_level"] = _classify_effort_hours(content_avg)

    # Link hours → off_page
    link_avg = _avg_nonzero(breakdown.get("link", []))
    per_focus["off_page"]["monthly_hours"] = round(link_avg, 2)
    per_focus["off_page"]["effort_level"] = _classify_effort_hours(link_avg)


# ── Content plan parser ───────────────────────────────────────────────────────


def _parse_content_plan(ws) -> list[dict]:
    """Parse the '4. Content' worksheet into a list of content item dicts.

    Header is at row 7; data starts at row 8.
    Columns (0-indexed): A=0 Month#, B=1 Month Name, C=2 URL, D=3 Title,
    E=4 Focus, F=5 Priority, G=6 Content Type, H=7 Word Count, I=8 SEO Hours
    """
    items = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        if not row:
            continue
        # URL is at col C (index 2) — skip empty rows
        url_val = row[2] if len(row) > 2 else None
        if url_val is None or str(url_val).strip() == "":
            continue

        month = int(row[0]) if row[0] and isinstance(row[0], (int, float)) else None
        month_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        url = str(url_val).strip()
        title = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        focus = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        priority = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        content_desc = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        try:
            word_count = int(row[7]) if len(row) > 7 and row[7] is not None else 0
        except (ValueError, TypeError):
            word_count = 0
        try:
            seo_hours = float(row[8]) if len(row) > 8 and row[8] is not None else 0.0
        except (ValueError, TypeError):
            seo_hours = 0.0

        content_desc_lower = content_desc.lower()
        is_new_page = "new page" in content_desc_lower
        is_faq = "faq" in content_desc_lower

        if is_new_page:
            content_type = "new_page"
        elif is_faq:
            content_type = "faq"
        else:
            content_type = "optimisation"

        items.append(
            {
                "month": month,
                "month_name": month_name,
                "url": url,
                "title": title,
                "focus": focus,
                "priority": priority,
                "content_type": content_type,
                "word_count": word_count,
                "seo_hours": seo_hours,
            }
        )
    return items


# ── Task sheet parser ─────────────────────────────────────────────────────────


def _parse_task_sheet(ws) -> list[dict]:
    """Parse a Consulting/Technical/Links worksheet (header row 1, data row 2+).

    Returns a list of dicts with keys: task, focus, occurrence, hours.
    """
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Skip fully-blank rows
        if all(v is None for v in row):
            continue
        task_val = row[0] if len(row) > 0 else None
        focus_val = row[1] if len(row) > 1 else None
        occ_val = row[2] if len(row) > 2 else None
        hrs_val = row[3] if len(row) > 3 else None
        try:
            hours = float(hrs_val) if hrs_val is not None else 0.0
        except (ValueError, TypeError):
            hours = 0.0
        tasks.append(
            {
                "task": str(task_val).strip() if task_val is not None else "",
                "focus": str(focus_val).strip() if focus_val is not None else "",
                "occurrence": str(occ_val).strip() if occ_val is not None else "",
                "hours": hours,
            }
        )
    return tasks


# ── Timeline finalisation ─────────────────────────────────────────────────────


def _finalise_timeline(bundle: dict) -> None:
    """Compute timeline metadata and store it in bundle['timeline']."""
    content_plan = bundle.get("content_plan", [])
    months_with_content = [
        item["month"]
        for item in content_plan
        if item.get("month") and isinstance(item["month"], (int, float))
    ]
    last_month = int(max(months_with_content)) if months_with_content else 12

    consulting_tasks = bundle.get("_consulting_tasks", [])
    has_strategy_review = any(
        "strategy review" in (t.get("task") or "").lower() for t in consulting_tasks
    )

    strategy_restart_month: int | None = None
    if last_month < 12 and has_strategy_review:
        strategy_restart_month = last_month + 1

    bundle["timeline"] = {
        "months_covered": 12,
        "strategy_restart_month": strategy_restart_month,
        "has_launch_dates": bool(months_with_content),
    }

    # Clean up the private key
    bundle.pop("_consulting_tasks", None)


# ── Global rollup finalisation ────────────────────────────────────────────────


def _finalise_global_rollup(bundle: dict) -> None:
    """Compute a global rollup across all per_focus entries."""
    per_focus = bundle.get("per_focus", {})

    total_monthly_hours = sum(f.get("monthly_hours", 0.0) for f in per_focus.values())

    # Max effort across all focus areas
    effort_order = ["light", "moderate", "aggressive"]
    effort_levels = [f.get("effort_level", "moderate") for f in per_focus.values() if f.get("monthly_hours", 0.0) > 0]
    if effort_levels:
        max_effort = max(effort_levels, key=lambda e: effort_order.index(e) if e in effort_order else 0)
    else:
        max_effort = "moderate"

    # Content cadence: number of content items per month from content_plan
    content_plan = bundle.get("content_plan", [])
    new_pages = [item for item in content_plan if item.get("content_type") == "new_page"]
    # Deduplicate by month to get posts per month average
    months_seen: dict[int, int] = {}
    for item in new_pages:
        m = item.get("month")
        if m:
            months_seen[m] = months_seen.get(m, 0) + 1
    content_cadence = round(sum(months_seen.values()) / len(months_seen)) if months_seen else 0

    # Maintenance coverage: fraction of portfolio covered by technical + on_page
    tech_hours = per_focus.get("technical", {}).get("monthly_hours", 0.0)
    on_page_hours = per_focus.get("on_page", {}).get("monthly_hours", 0.0)
    maintenance_hours = tech_hours + on_page_hours
    maintenance_coverage = round(min(maintenance_hours / max(total_monthly_hours, 1.0), 1.0), 2)

    bundle["global_rollup"] = {
        "total_monthly_hours": round(total_monthly_hours, 2),
        "effort_level": max_effort,
        "maintenance_coverage": maintenance_coverage,
        "content_cadence": content_cadence,
        "positional_effort_level": per_focus.get("on_page", {}).get("effort_level", "moderate"),
    }


# ── AI extraction helpers ─────────────────────────────────────────────────────


def _parse_tasks_with_ai(ai_client, wb) -> tuple[dict, str]:
    """Extract tasks from Consulting/Technical/Links sheets via AI."""
    from engine.ai_engine import (
        _load_prompt,
        _parse_llm_json,
        generate_with_fallback,
        get_model_for_feature,
    )

    consulting_md = _ws_to_markdown(wb["2. Consulting"])[0] if "2. Consulting" in wb.sheetnames else "(sheet not found)"
    technical_md = _ws_to_markdown(wb["3. Technical"])[0] if "3. Technical" in wb.sheetnames else "(sheet not found)"
    links_md = _ws_to_markdown(wb["5. Links"])[0] if "5. Links" in wb.sheetnames else "(sheet not found)"

    system, user_tmpl = _load_prompt("extract_roadmap_tasks")
    user_input = user_tmpl.substitute(
        consulting_markdown=consulting_md,
        technical_markdown=technical_md,
        links_markdown=links_md,
    )
    model = get_model_for_feature("roadmap_extraction")
    text, used_model = generate_with_fallback(
        ai_client, model, system, user_input, temperature=0.1, max_tokens=4000,
    )
    return _parse_llm_json(text), used_model


def _parse_content_plan_with_ai(ai_client, wb) -> tuple[dict, str]:
    """Extract content plan from the Content sheet via AI."""
    from engine.ai_engine import (
        _load_prompt,
        _parse_llm_json,
        generate_with_fallback,
        get_model_for_feature,
    )

    content_md = _ws_to_markdown(wb["4. Content"], max_chars=18000)[0] if "4. Content" in wb.sheetnames else "(sheet not found)"

    system, user_tmpl = _load_prompt("extract_roadmap_content_plan")
    user_input = user_tmpl.substitute(content_markdown=content_md)
    model = get_model_for_feature("roadmap_extraction")
    text, used_model = generate_with_fallback(
        ai_client, model, system, user_input, temperature=0.1, max_tokens=8000,
    )
    return _parse_llm_json(text), used_model


def _ai_tasks_to_per_focus(tasks_result: dict, breakdown_per_focus: dict) -> dict:
    """Merge AI-extracted tasks into per_focus, preserving breakdown-derived hours."""
    per_focus = {k: {**v, "tasks": list(v.get("tasks", []))} for k, v in breakdown_per_focus.items()}

    all_tasks = (
        tasks_result.get("consulting_tasks", [])
        + tasks_result.get("technical_tasks", [])
        + tasks_result.get("links_tasks", [])
    )

    _focus_aliases: dict[str, str] = {
        "off-page": "off_page",
        "on-page": "on_page",
    }

    for task in all_tasks:
        raw_focus = (task.get("focus") or "strategy").lower().replace(" ", "_")
        focus = _focus_aliases.get(raw_focus, raw_focus)
        if focus not in per_focus:
            focus = "strategy"

        per_focus[focus]["tasks"].append({
            "name": str(task.get("task") or ""),
            "hours": _safe_float(task.get("hours")),
            "occurrence": str(task.get("occurrence") or "monthly"),
            "contribution": "primary",
        })
        per_focus[focus]["task_count"] = len(per_focus[focus]["tasks"])

    return per_focus


def _validate_bundle(bundle: dict, source_file: str) -> None:
    """Validate bundle completeness. Raises ValueError listing all errors."""
    errors: list[str] = []

    if bundle.get("schema_version") != "2.0":
        errors.append("schema_version must be '2.0'")

    per_focus = bundle.get("per_focus") or {}
    missing = {"content", "technical", "on_page", "off_page", "local", "analytics", "strategy"} - set(per_focus.keys())
    if missing:
        errors.append(f"per_focus missing keys: {sorted(missing)}")

    if bundle.get("content_plan") is None:
        errors.append("content_plan missing from bundle")

    if "total_monthly_hours" not in (bundle.get("global_rollup") or {}):
        errors.append("global_rollup.total_monthly_hours missing")

    if errors:
        raise ValueError(
            f"Roadmap bundle validation failed ({source_file}):\n"
            + "\n".join(f"  - {e}" for e in errors)
        )


# ── Main parser ───────────────────────────────────────────────────────────────


def parse_pattern_native(
    raw_bytes: bytes,
    ai_client=None,
    source_filename: str = "roadmap.xlsx",
) -> dict:
    """Parse a Pattern-native xlsx roadmap into a v2 bundle dict.

    When ai_client is provided, uses AI extraction for variable-layout sheets
    (Consulting, Technical, Content, Links). Client Detail and Breakdown are
    always parsed deterministically.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    sheet_names = wb.sheetnames

    bundle: dict = {
        "schema_version": "2.0",
        "format_detected": "pattern_native",
        "extraction_method": "hybrid_ai" if ai_client is not None else "deterministic",
        "source_summary": {
            "sheets_parsed": [],
            "parsing_confidence": 0.95 if ai_client is not None else 0.80,
        },
        "client_metadata": {},
        "per_focus": _empty_per_focus(),
        "content_plan": [],
        "timeline": {},
        "global_rollup": {},
        "recommendations": [],
        "gaps": [],
        "strategy_summary": "",
        "primary_domain": "",
        "localisation_domains": [],
    }

    # ── Deterministic: Client Detail ──────────────────────────────────────────
    if "1. Client Detail" in sheet_names:
        bundle["client_metadata"] = _parse_client_detail(wb["1. Client Detail"])
        bundle["source_summary"]["sheets_parsed"].append("1. Client Detail")

    # ── Deterministic: Breakdown ──────────────────────────────────────────────
    if "Breakdown" in sheet_names:
        breakdown = _parse_breakdown(wb["Breakdown"])
        _apply_breakdown_to_bundle(bundle, breakdown)
        bundle["source_summary"]["sheets_parsed"].append("Breakdown")

    # ── AI path or deterministic fallback ─────────────────────────────────────
    if ai_client is not None:
        try:
            tasks_result, _ = _parse_tasks_with_ai(ai_client, wb)
            bundle["per_focus"] = _ai_tasks_to_per_focus(tasks_result, bundle["per_focus"])
            for sh in ("2. Consulting", "3. Technical", "5. Links"):
                if sh in sheet_names:
                    bundle["source_summary"]["sheets_parsed"].append(sh)
        except Exception as exc:
            bundle["recommendations"].append({
                "severity": "warning",
                "message": f"AI task extraction failed ({exc}); task lists unavailable",
            })

        if "4. Content" in sheet_names:
            try:
                content_result, _ = _parse_content_plan_with_ai(ai_client, wb)
                bundle["content_plan"] = content_result.get("content_plan", [])
                bundle["strategy_summary"] = content_result.get("strategy_summary", "")
                bundle["primary_domain"] = content_result.get("primary_domain", "")
                bundle["localisation_domains"] = content_result.get("localisation_domains", [])
                bundle["source_summary"]["sheets_parsed"].append("4. Content")
                bundle["source_summary"]["content_launches_detected"] = len(bundle["content_plan"])
            except Exception as exc:
                bundle["recommendations"].append({
                    "severity": "warning",
                    "message": f"AI content extraction failed ({exc}); falling back to deterministic",
                })
                bundle["content_plan"] = _parse_content_plan(wb["4. Content"])
                bundle["source_summary"]["sheets_parsed"].append("4. Content")
                bundle["source_summary"]["content_launches_detected"] = len(bundle["content_plan"])
    else:
        bundle["recommendations"].append({
            "severity": "info",
            "message": (
                "No AI client provided. Provide a Bi Frost API key for richer task "
                "and content extraction."
            ),
        })
        if "4. Content" in sheet_names:
            bundle["content_plan"] = _parse_content_plan(wb["4. Content"])
            bundle["source_summary"]["content_launches_detected"] = len(bundle["content_plan"])
            bundle["source_summary"]["sheets_parsed"].append("4. Content")

        if "2. Consulting" in sheet_names:
            bundle["_consulting_tasks"] = _parse_task_sheet(wb["2. Consulting"])
            bundle["source_summary"]["sheets_parsed"].append("2. Consulting")

        if "3. Technical" in sheet_names:
            _parse_task_sheet(wb["3. Technical"])
            bundle["source_summary"]["sheets_parsed"].append("3. Technical")

        if "5. Links" in sheet_names:
            _parse_task_sheet(wb["5. Links"])
            bundle["source_summary"]["sheets_parsed"].append("5. Links")

    _finalise_timeline(bundle)
    _finalise_global_rollup(bundle)
    _validate_bundle(bundle, source_filename)
    return bundle


# ── Legacy wrappers ───────────────────────────────────────────────────────────


def wrap_legacy_task_table_as_bundle(legacy_result: dict) -> dict:
    """Wrap a legacy parse_task_table() output dict into a v2 bundle.

    Args:
        legacy_result: Dict with keys: content_cadence, effort_level,
            maintenance_coverage (and optional _monthly_hours).

    Returns:
        A v2 bundle with parsing_confidence = 0.7.
    """
    effort = legacy_result.get("effort_level", "moderate")
    cadence = legacy_result.get("content_cadence", 4)
    maintenance = legacy_result.get("maintenance_coverage", 0.0)
    monthly_hours = legacy_result.get("_monthly_hours", 0.0)

    per_focus = _empty_per_focus()
    # Apply effort to content, on_page, off_page (the focus areas that the
    # task-table loader addresses)
    for area in ("content", "on_page", "off_page"):
        per_focus[area]["effort_level"] = effort

    bundle: dict = {
        "schema_version": "2.0",
        "format_detected": "task_table",
        "extraction_method": "deterministic",
        "source_summary": {
            "sheets_parsed": [],
            "parsing_confidence": 0.7,
        },
        "client_metadata": {},
        "per_focus": per_focus,
        "content_plan": [],
        "timeline": {
            "months_covered": 12,
            "strategy_restart_month": None,
            "has_launch_dates": False,
        },
        "global_rollup": {
            "total_monthly_hours": float(monthly_hours),
            "effort_level": effort,
            "maintenance_coverage": float(maintenance),
            "content_cadence": int(cadence),
            "positional_effort_level": effort,
        },
        "recommendations": [],
        "gaps": [],
    }
    return bundle


def wrap_legacy_param_table_as_bundle(legacy_result: dict) -> dict:
    """Wrap a legacy parse_param_table() output dict into a v2 bundle.

    Args:
        legacy_result: Dict with optional keys: content_cadence, effort_level,
            maintenance_coverage.

    Returns:
        A v2 bundle with parsing_confidence = 0.5.
    """
    effort = legacy_result.get("effort_level", "moderate")
    cadence = legacy_result.get("content_cadence", 4)
    maintenance = legacy_result.get("maintenance_coverage", 0.0)

    per_focus = _empty_per_focus()

    bundle: dict = {
        "schema_version": "2.0",
        "format_detected": "param_table",
        "extraction_method": "deterministic",
        "source_summary": {
            "sheets_parsed": [],
            "parsing_confidence": 0.5,
        },
        "client_metadata": {},
        "per_focus": per_focus,
        "content_plan": [],
        "timeline": {
            "months_covered": 12,
            "strategy_restart_month": None,
            "has_launch_dates": False,
        },
        "global_rollup": {
            "total_monthly_hours": 0.0,
            "effort_level": effort,
            "maintenance_coverage": float(maintenance),
            "content_cadence": int(cadence),
            "positional_effort_level": effort,
        },
        "recommendations": [],
        "gaps": [],
    }
    return bundle
